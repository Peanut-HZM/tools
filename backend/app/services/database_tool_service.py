import uuid
import json
import logging
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Optional, Dict, Any
from datetime import datetime

from app.config.database import get_pooled_db_connection, release_db_connection
from app.utils.db_error_mapper import map_connection_error
from app.models.database_tool_models import (
    DatabaseConfigBase,
    CreateDatabaseRequest,
    UpdateDatabaseRequest,
    DatabaseConfigResponse,
    TestConnectionRequest,
    ConnectionTestResult,
    SQLExecutionRequest,
    SQLExecutionResult,
    ExecutionHistory,
    TableSchema,
    TableData,
    DatabaseType,
    ColumnDefinition,
    TableModificationRequest,
    ExportDataRequest,
    ExportDataResponse,
    ExportFormat,
    ImportDataRequest,
    ImportDataResponse,
    ExplainPlanRequest,
    ExplainPlanResponse,
    ExplainPlanStep,
    TablePreviewRequest,
    TablePreviewResponse,
    AutoCompleteRequest,
    AutoCompleteResponse,
    AutoCompleteItem,
    BackupDatabaseRequest,
    BackupDatabaseResponse,
    BackupRecordResponse,
    RestoreDatabaseRequest,
    RestoreDatabaseResponse,
    TableDetailResponse,
    ColumnDetail,
    IndexDetail,
    ForeignKeyDetail,
    DisplayPreference,
    DisplayPreferenceResponse,
)
from app.services.backup_generators import get_generator
from app.services.backup_storage import backup_storage
from app.utils.encryption import EncryptionUtils
from app.utils.db_connection_manager import DBConnectionManager
from app.utils.sql_executor import SQLExecutor, sqlparse, is_executable_statement
from sqlalchemy import inspect, text

logger = logging.getLogger(__name__)


class StructureCache:
    """线程安全的 TTL 缓存，无需外部依赖"""
    def __init__(self, ttl: int = 600, maxsize: int = 100):
        self._cache: Dict[str, Any] = {}
        self._timestamps: Dict[str, float] = {}
        self._ttl = ttl
        self._maxsize = maxsize
        self._lock = threading.Lock()

    def get(self, key: str) -> Optional[Any]:
        with self._lock:
            if key in self._cache:
                if time.time() - self._timestamps[key] < self._ttl:
                    return self._cache[key]
                else:
                    del self._cache[key]
                    del self._timestamps[key]
            return None

    def set(self, key: str, value: Any) -> None:
        with self._lock:
            if len(self._cache) >= self._maxsize and key not in self._cache:
                oldest_key = min(self._timestamps, key=self._timestamps.get)
                del self._cache[oldest_key]
                del self._timestamps[oldest_key]
            self._cache[key] = value
            self._timestamps[key] = time.time()

    def invalidate(self, key: str) -> None:
        with self._lock:
            self._cache.pop(key, None)
            self._timestamps.pop(key, None)

    def invalidate_prefix(self, prefix: str) -> None:
        with self._lock:
            keys_to_remove = [k for k in list(self._cache.keys()) if k.startswith(prefix)]
            for k in keys_to_remove:
                del self._cache[k]
                del self._timestamps[k]


# 全局实例：1 小时 TTL，最多 100 条
_STRUCTURE_CACHE = StructureCache(ttl=3600, maxsize=100)

# 库名/Schema 列表缓存：5 分钟 TTL（schema 变动不频繁，靠右键刷新与写操作失效兜底）
_LIST_CACHE = StructureCache(ttl=300, maxsize=200)


class DatabaseToolService:
    @staticmethod
    def _decrypt_password(password_encrypted: str, config_id: str = "") -> tuple[str | None, str | None]:
        """
        解密数据库配置密码。支持多密钥自动迁移：
        1. 优先用数据库存储的密钥解密
        2. 失败则尝试 .env 中的密钥
        3. 再失败则尝试默认开发密钥
        成功则自动用新密钥重新加密并更新数据库
        返回 (password, error_message)。解密失败时返回 (None, 错误描述)，不会抛异常。
        """
        if not password_encrypted:
            return None, "数据库配置密码为空"

        def _try_decrypt_with_key(key: str) -> str | None:
            try:
                from cryptography.fernet import Fernet
                import base64 as _base64, hashlib as _hashlib
                try:
                    suite = Fernet(key)
                except ValueError:
                    m = _hashlib.sha256()
                    m.update(key.encode("utf-8"))
                    suite = Fernet(_base64.urlsafe_b64encode(m.digest()))
                return suite.decrypt(password_encrypted.encode("utf-8")).decode("utf-8")
            except Exception:
                return None

        # 优先使用数据库中的密钥
        try:
            password = EncryptionUtils.decrypt(password_encrypted)
            return password, None
        except Exception:
            pass

        # 尝试各种旧密钥
        from app.config.config import settings as _settings
        fallback_keys = [
            _settings.DB_ENCRYPTION_KEY,
            "dev-db-encryption-change-me",
        ]
        for key in fallback_keys:
            password = _try_decrypt_with_key(key)
            if password is not None:
                # 用新密钥重新加密并更新数据库
                try:
                    new_encrypted = EncryptionUtils.encrypt(password)
                    DatabaseToolService._re_encrypt_password(config_id, new_encrypted)
                    logger.info(f"配置 {config_id} 密码已从旧密钥迁移到新密钥")
                except Exception as e:
                    logger.warning(f"配置 {config_id} 密码迁移到新密钥失败: {e}")
                return password, None

        logger.error(f"数据库配置 {config_id} 密码解密失败（所有密钥尝试均失败）")
        return None, "密码解密失败，请编辑该连接重新保存密码"

    @staticmethod
    def _re_encrypt_password(config_id: str, new_encrypted: str):
        """用新密钥更新数据库中存储的加密密码"""
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    "UPDATE db_configs SET password_encrypted = %s, updated_at = NOW() WHERE id = %s",
                    (new_encrypted, config_id),
                )
                conn.commit()
        except Exception as e:
            logger.error(f"更新配置 {config_id} 加密密码失败: {e}")
        finally:
            release_db_connection(conn)

    @staticmethod
    def generate_ddl(
        user_id: str, config_id: str, table_name: str, database_name: str, schema_name: str = None
    ) -> str:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine = DBConnectionManager.get_engine(config_id, config_dict)
        db_type = config_row["db_type"]

        try:
            if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                with engine.connect() as conn:
                    result = conn.execute(text(f"SHOW CREATE TABLE `{table_name}`"))
                    row = result.fetchone()
                    if row and len(row) > 1:
                        return row[1]
            elif db_type == DatabaseType.SQLITE:
                with engine.connect() as conn:
                    result = conn.execute(
                        text(
                            f"SELECT sql FROM sqlite_master WHERE type='table' AND name='{table_name}'"
                        )
                    )
                    row = result.fetchone()
                    if row:
                        return row[0]
            elif db_type == DatabaseType.POSTGRESQL:
                with engine.connect() as conn:
                    schema_filter = f"AND n.nspname = '{schema_name}'" if schema_name else ""
                    result = conn.execute(text(f"""
                        SELECT pg_catalog.pg_get_tabledef(c.oid)
                        FROM pg_class c
                        LEFT JOIN pg_namespace n ON n.oid = c.relnamespace
                        WHERE c.relname = '{table_name}' {schema_filter}
                        AND c.relkind IN ('r', 'p')
                        LIMIT 1
                    """))
                    row = result.fetchone()
                    if row and row[0]:
                        return row[0]
                    # Fallback: get DDL from information_schema
                    result = conn.execute(text(f"""
                        SELECT column_name, data_type, is_nullable, column_default
                        FROM information_schema.columns
                        WHERE table_name = '{table_name}' {schema_filter}
                        ORDER BY ordinal_position
                    """))
                    rows = result.fetchall()
                    if rows:
                        cols = []
                        for r in rows:
                            col_def = f"    {r[0]} {r[1]}"
                            if r[2] == 'NO':
                                col_def += " NOT NULL"
                            if r[3]:
                                col_def += f" DEFAULT {r[3]}"
                            cols.append(col_def)
                        schema_part = f'"{schema_name}".' if schema_name else ''
                        return f'CREATE TABLE {schema_part}"{table_name}" (\n{",\n".join(cols)}\n);'

            # Fallback for others using SQLAlchemy
            from sqlalchemy.schema import CreateTable
            from sqlalchemy import MetaData, Table

            metadata = MetaData()
            if db_type == DatabaseType.POSTGRESQL and schema_name:
                table = Table(table_name, metadata, autoload_with=engine, schema=schema_name)
            else:
                table = Table(table_name, metadata, autoload_with=engine)
            return str(CreateTable(table).compile(engine))

        except Exception as e:
            logger.error(f"Failed to generate DDL: {e}")
            raise e

    @staticmethod
    def modify_table_structure(
        user_id: str, config_id: str, request: TableModificationRequest
    ) -> bool:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{request.database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        db_type = config_row["db_type"]

        # Schema-qualified table name for PostgreSQL
        if db_type == DatabaseType.POSTGRESQL and request.schema_name:
            qualified_table = f'"{request.schema_name}"."{request.table_name}"'
        elif db_type in (DatabaseType.MYSQL, DatabaseType.MARIADB):
            qualified_table = f"`{request.table_name}`"
        else:
            qualified_table = f"`{request.table_name}`"

        try:
            inspector = inspect(engine)
            # For PostgreSQL with schema, use schema-qualified name
            if db_type == DatabaseType.POSTGRESQL and request.schema_name:
                existing_columns = inspector.get_columns(qualified_table, schema=request.schema_name)
            else:
                existing_columns = inspector.get_columns(request.table_name)
            existing_col_map = {col["name"]: col for col in existing_columns}

            statements = []

            # 1. Handle Columns
            # We need to preserve order, but ALTER usually appends or needs AFTER/FIRST
            # For simplicity, we just generate ADD/MODIFY/DROP. Reordering is hard.

            # Helper for MySQL type mapping/formatting
            def _escape_sql_string(value: str) -> str:
                """转义 SQL 字符串中的单引号，防止 SQL 注入和语法错误"""
                return value.replace("'", "\\'")

            def _strip_outer_quotes(value: str) -> str:
                """去除字符串外层包裹的单引号或双引号"""
                value = value.strip()
                if len(value) >= 2:
                    if (value[0] == "'" and value[-1] == "'") or (
                        value[0] == '"' and value[-1] == '"'
                    ):
                        return value[1:-1]
                return value

            def format_col_def(col: ColumnDefinition):
                def_str = f"`{col.name}` {col.type}"
                if col.length:
                    def_str += f"({col.length})"

                if not col.nullable:
                    def_str += " NOT NULL"
                else:
                    def_str += " NULL"

                if col.default_value is not None:
                    # 清理默认值中的外层引号，避免嵌套引号导致语法错误
                    default_val = _strip_outer_quotes(str(col.default_value)).strip()
                    if default_val.upper() in ["CURRENT_TIMESTAMP", "NULL"]:
                        def_str += f" DEFAULT {default_val}"
                    else:
                        # 使用单引号包裹，并转义内部单引号
                        def_str += f" DEFAULT '{_escape_sql_string(default_val)}'"

                if col.auto_increment:
                    def_str += " AUTO_INCREMENT"

                if col.comment:
                    # 转义注释中的单引号，防止 SQL 语法错误
                    def_str += f" COMMENT '{_escape_sql_string(col.comment)}'"

                return def_str

            new_col_names = set()

            for col in request.columns:
                new_col_names.add(col.name)
                if col.name not in existing_col_map:
                    # ADD
                    statements.append(
                        f"ALTER TABLE {qualified_table} ADD COLUMN {format_col_def(col)}"
                    )
                else:
                    # MODIFY
                    statements.append(
                        f"ALTER TABLE {qualified_table} MODIFY COLUMN {format_col_def(col)}"
                    )

            # DROP columns
            for existing_name in existing_col_map:
                if existing_name not in new_col_names:
                    statements.append(
                        f"ALTER TABLE {qualified_table} DROP COLUMN `{existing_name}`"
                    )

            # 2. Handle Table Comment
            if request.comment is not None:
                statements.append(
                    f"ALTER TABLE {qualified_table} COMMENT = '{_escape_sql_string(request.comment)}'"
                )

            # 3. Handle Table Rename
            if request.new_table_name and request.new_table_name != request.table_name:
                statements.append(
                    f"ALTER TABLE {qualified_table} RENAME TO `{request.new_table_name}`"
                )

            # Execute all statements
            with engine.connect().execution_options(
                isolation_level="AUTOCOMMIT"
            ) as conn:
                for sql in statements:
                    conn.execute(text(sql))

            return True

        except Exception as e:
            logger.error(f"Failed to modify table: {e}")
            raise e

    @staticmethod
    def delete_all_tables(user_id: str, config_id: str, database_name: str) -> bool:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        db_type = config_row["db_type"]

        try:
            with engine.connect().execution_options(
                isolation_level="AUTOCOMMIT"
            ) as conn:
                # Disable foreign key checks for MySQL to avoid constraint errors
                if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                    conn.execute(text("SET FOREIGN_KEY_CHECKS = 0"))

                inspector = inspect(engine)
                tables = inspector.get_table_names()

                for table in tables:
                    if db_type == DatabaseType.POSTGRESQL:
                        conn.execute(text(f'DROP TABLE IF EXISTS "{table}" CASCADE'))
                    else:
                        conn.execute(text(f"DROP TABLE IF EXISTS `{table}`"))

                # Re-enable foreign key checks
                if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                    conn.execute(text("SET FOREIGN_KEY_CHECKS = 1"))

            return True
        except Exception as e:
            logger.error(f"Failed to delete all tables: {e}")
            raise e

    @staticmethod
    def truncate_all_tables(user_id: str, config_id: str, database_name: str) -> bool:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        db_type = config_row["db_type"]

        try:
            with engine.connect().execution_options(
                isolation_level="AUTOCOMMIT"
            ) as conn:
                # Disable foreign key checks for MySQL
                if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                    conn.execute(text("SET FOREIGN_KEY_CHECKS = 0"))

                inspector = inspect(engine)
                tables = inspector.get_table_names()

                for table in tables:
                    if db_type == DatabaseType.SQLITE:
                        conn.execute(text(f"DELETE FROM `{table}`"))
                    elif db_type == DatabaseType.POSTGRESQL:
                        conn.execute(text(f'TRUNCATE TABLE "{table}" CASCADE'))
                    else:
                        conn.execute(text(f"TRUNCATE TABLE `{table}`"))

                # Re-enable foreign key checks
                if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                    conn.execute(text("SET FOREIGN_KEY_CHECKS = 1"))

            return True
        except Exception as e:
            logger.error(f"Failed to truncate all tables: {e}")
            raise e

    @staticmethod
    def generate_all_tables_ddl(
        user_id: str, config_id: str, database_name: str
    ) -> str:
        # Re-use existing generate_ddl logic but for all tables
        # Since getting all DDLs might be heavy, we should do it sequentially
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        # We can reuse the single table generation logic
        # But let's avoid fetching config every time

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)

        inspector = inspect(engine)
        tables = inspector.get_table_names()

        full_ddl = ""

        for table in tables:
            try:
                # We can call the static method if we want, but better to inline or call helper to avoid overhead
                # Let's call helper
                ddl = DatabaseToolService.generate_ddl(
                    user_id, config_id, table, database_name
                )
                full_ddl += f"-- Table: {table}\n{ddl};\n\n"
            except Exception as e:
                full_ddl += f"-- Failed to generate DDL for {table}: {str(e)}\n\n"

        return full_ddl

    # --------------------------------------------------------------------------
    # Database Configuration Management
    # --------------------------------------------------------------------------

    @staticmethod
    def get_all_configs(
        user_id: str, include_password: bool = False
    ) -> List[DatabaseConfigResponse]:
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    "SELECT * FROM db_configs WHERE user_id = %s AND deleted = FALSE ORDER BY created_at DESC",
                    (user_id,),
                )
                rows = cursor.fetchall()

                configs = []
                for row in rows:
                    password = None
                    if include_password:
                        password, _ = DatabaseToolService._decrypt_password(
                            row["password_encrypted"], row["id"]
                        )
                    # 'row' is a RealDictRow
                    config = DatabaseConfigResponse(
                        id=row["id"],
                        user_id=row["user_id"],
                        alias=row["alias"],
                        db_type=row["db_type"],
                        host=row["host"],
                        port=row["port"],
                        database_name=row["database_name"],
                        username=row["username"],
                        environment=row["environment"],
                        group_name=row["group_name"],
                        charset=row["charset"],
                        connect_timeout=row["connect_timeout"],
                        max_pool_size=row["max_pool_size"],
                        ssl_mode=row["ssl_mode"],
                        ssl_cert_path=row["ssl_cert_path"],
                        extra_config=row["extra_config"]
                        if isinstance(row["extra_config"], dict)
                        else (
                            json.loads(row["extra_config"])
                            if row["extra_config"]
                            else None
                        ),
                        is_active=row["is_active"],
                        last_connected_at=row["last_connected_at"],
                        created_at=row["created_at"],
                        updated_at=row["updated_at"],
                        password=password,
                    )
                    configs.append(config)
                return configs
        finally:
            release_db_connection(conn)

    @staticmethod
    def get_config(
        config_id: str, user_id: str, include_password: bool = False
    ) -> Optional[DatabaseConfigResponse]:
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    "SELECT * FROM db_configs WHERE id = %s AND user_id = %s AND deleted = FALSE",
                    (config_id, user_id),
                )
                row = cursor.fetchone()
                if not row:
                    return None
                password = None
                if include_password:
                    try:
                        password = EncryptionUtils.decrypt(row["password_encrypted"])
                    except Exception:
                        password = None
                return DatabaseConfigResponse(
                    id=row["id"],
                    user_id=row["user_id"],
                    alias=row["alias"],
                    db_type=row["db_type"],
                    host=row["host"],
                    port=row["port"],
                    database_name=row["database_name"],
                    username=row["username"],
                    environment=row["environment"],
                    group_name=row["group_name"],
                    charset=row["charset"],
                    connect_timeout=row["connect_timeout"],
                    max_pool_size=row["max_pool_size"],
                    ssl_mode=row["ssl_mode"],
                    ssl_cert_path=row["ssl_cert_path"],
                    extra_config=row["extra_config"]
                    if isinstance(row["extra_config"], dict)
                    else (
                        json.loads(row["extra_config"]) if row["extra_config"] else None
                    ),
                    is_active=row["is_active"],
                    last_connected_at=row["last_connected_at"],
                    created_at=row["created_at"],
                    updated_at=row["updated_at"],
                    password=password,
                )
        finally:
            release_db_connection(conn)

    @staticmethod
    def _get_config_with_password(
        config_id: str, user_id: str
    ) -> Optional[Dict[str, Any]]:
        """Internal helper to get config including encrypted password"""
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    "SELECT * FROM db_configs WHERE id = %s AND user_id = %s AND deleted = FALSE",
                    (config_id, user_id),
                )
                row = cursor.fetchone()
                if not row:
                    return None
                return dict(row)
        finally:
            release_db_connection(conn)

    @staticmethod
    def create_config(
        user_id: str, request: CreateDatabaseRequest
    ) -> DatabaseConfigResponse:
        conn = get_pooled_db_connection()
        try:
            # Encrypt password
            encrypted_password = EncryptionUtils.encrypt(request.password)

            config_id = str(uuid.uuid4())
            now = datetime.now()

            with conn.cursor() as cursor:
                # Check alias uniqueness for user
                cursor.execute(
                    "SELECT id FROM db_configs WHERE user_id = %s AND alias = %s AND deleted = FALSE",
                    (user_id, request.alias),
                )
                if cursor.fetchone():
                    raise ValueError(f"Alias '{request.alias}' already exists")

                cursor.execute(
                    """
                    INSERT INTO db_configs (
                        id, user_id, alias, db_type, host, port, database_name, username, 
                        password_encrypted, environment, group_name, charset, connect_timeout, 
                        max_pool_size, ssl_mode, ssl_cert_path, extra_config, is_active, 
                        created_at, updated_at, deleted
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, 
                        %s, %s, %s, %s, %s, 
                        %s, %s, %s, %s, %s, 
                        %s, %s, FALSE
                    )
                    """,
                    (
                        config_id,
                        user_id,
                        request.alias,
                        request.db_type,
                        request.host,
                        request.port,
                        request.database_name,
                        request.username,
                        encrypted_password,
                        request.environment,
                        request.group_name,
                        request.charset,
                        request.connect_timeout,
                        request.max_pool_size,
                        request.ssl_mode,
                        request.ssl_cert_path,
                        json.dumps(request.extra_config)
                        if request.extra_config
                        else None,
                        request.is_active,
                        now,
                        now,
                    ),
                )
                conn.commit()

            return DatabaseToolService.get_config(config_id, user_id)
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            release_db_connection(conn)

    @staticmethod
    def update_config(
        config_id: str, user_id: str, request: UpdateDatabaseRequest
    ) -> Optional[DatabaseConfigResponse]:
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                # Check existence
                cursor.execute(
                    "SELECT id FROM db_configs WHERE id = %s AND user_id = %s AND deleted = FALSE",
                    (config_id, user_id),
                )
                if not cursor.fetchone():
                    return None

                updates = []
                params = []

                if request.alias is not None:
                    # Check alias uniqueness if changed
                    cursor.execute(
                        "SELECT id FROM db_configs WHERE user_id = %s AND alias = %s AND id != %s AND deleted = FALSE",
                        (user_id, request.alias, config_id),
                    )
                    if cursor.fetchone():
                        raise ValueError(f"Alias '{request.alias}' already exists")
                    updates.append("alias = %s")
                    params.append(request.alias)

                if request.host is not None:
                    updates.append("host = %s")
                    params.append(request.host)
                if request.port is not None:
                    updates.append("port = %s")
                    params.append(request.port)
                if request.database_name is not None:
                    updates.append("database_name = %s")
                    params.append(request.database_name)
                if request.username is not None:
                    updates.append("username = %s")
                    params.append(request.username)
                if request.password is not None:
                    encrypted = EncryptionUtils.encrypt(request.password)
                    updates.append("password_encrypted = %s")
                    params.append(encrypted)
                if request.environment is not None:
                    updates.append("environment = %s")
                    params.append(request.environment)
                if request.group_name is not None:
                    updates.append("group_name = %s")
                    params.append(request.group_name)
                if request.charset is not None:
                    updates.append("charset = %s")
                    params.append(request.charset)
                if request.connect_timeout is not None:
                    updates.append("connect_timeout = %s")
                    params.append(request.connect_timeout)
                if request.max_pool_size is not None:
                    updates.append("max_pool_size = %s")
                    params.append(request.max_pool_size)
                if request.ssl_mode is not None:
                    updates.append("ssl_mode = %s")
                    params.append(request.ssl_mode)
                if request.ssl_cert_path is not None:
                    updates.append("ssl_cert_path = %s")
                    params.append(request.ssl_cert_path)
                if request.extra_config is not None:
                    updates.append("extra_config = %s")
                    params.append(json.dumps(request.extra_config))
                if request.is_active is not None:
                    updates.append("is_active = %s")
                    params.append(request.is_active)

                if not updates:
                    return DatabaseToolService.get_config(config_id, user_id)

                updates.append("updated_at = %s")
                params.append(datetime.now())

                params.append(config_id)
                params.append(user_id)

                sql = f"UPDATE db_configs SET {', '.join(updates)} WHERE id = %s AND user_id = %s"
                cursor.execute(sql, tuple(params))
                conn.commit()

            # Clear engine cache if connection params changed
            DBConnectionManager.close_engine(config_id)

            return DatabaseToolService.get_config(config_id, user_id)
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            release_db_connection(conn)

    @staticmethod
    def delete_config(config_id: str, user_id: str) -> bool:
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                # Soft delete
                cursor.execute(
                    "UPDATE db_configs SET deleted = TRUE, updated_at = %s WHERE id = %s AND user_id = %s",
                    (datetime.now(), config_id, user_id),
                )
                if cursor.rowcount == 0:
                    return False
                conn.commit()

            # Close engine
            DBConnectionManager.close_engine(config_id)
            return True
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            release_db_connection(conn)

    @staticmethod
    def test_connection(request: TestConnectionRequest) -> ConnectionTestResult:
        config_dict = {
            "db_type": request.db_type,
            "host": request.host,
            "port": request.port,
            "database_name": request.database_name,
            "username": request.username,
            "password": request.password,
            "ssl_mode": request.ssl_mode,
            "ssl_cert_path": request.ssl_cert_path,
        }

        start_time = datetime.now()
        try:
            DBConnectionManager.test_connection(config_dict)
            elapsed = (datetime.now() - start_time).total_seconds() * 1000
            return ConnectionTestResult(
                success=True, message="Connection successful", elapsed_ms=elapsed
            )
        except Exception as e:
            elapsed = (datetime.now() - start_time).total_seconds() * 1000
            error_code, zh_msg = map_connection_error(str(e))
            return ConnectionTestResult(
                success=False,
                message=zh_msg,
                error_code=error_code,
                elapsed_ms=elapsed,
            )

    @staticmethod
    def test_connection_by_id(config_id: str, user_id: str) -> ConnectionTestResult:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            return ConnectionTestResult(
                success=False, message="Configuration not found"
            )

        # Decrypt password
        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            return ConnectionTestResult(
                success=False, message="Failed to decrypt password"
            )

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "ssl_mode": config_row["ssl_mode"],
            "ssl_cert_path": config_row["ssl_cert_path"],
            "charset": config_row["charset"],
        }

        start_time = datetime.now()
        try:
            DBConnectionManager.test_connection(config_dict)
            elapsed = (datetime.now() - start_time).total_seconds() * 1000

            # Update last_connected_at
            conn = get_pooled_db_connection()
            try:
                with conn.cursor() as cursor:
                    cursor.execute(
                        "UPDATE db_configs SET last_connected_at = %s WHERE id = %s",
                        (datetime.now(), config_id),
                    )
                    conn.commit()
            except:
                pass  # Ignore update failure
            finally:
                release_db_connection(conn)

            return ConnectionTestResult(
                success=True, message="Connection successful", elapsed_ms=elapsed
            )
        except Exception as e:
            elapsed = (datetime.now() - start_time).total_seconds() * 1000
            error_code, zh_msg = map_connection_error(str(e))
            return ConnectionTestResult(
                success=False,
                message=zh_msg,
                error_code=error_code,
                elapsed_ms=elapsed,
            )

    # --------------------------------------------------------------------------
    # SQL Execution
    # --------------------------------------------------------------------------

    @staticmethod
    def _apply_pagination(sql: str, page: int, page_size: int, db_type: str) -> str:
        # Simple check if pagination already exists
        sql_upper = sql.upper()
        if "LIMIT" in sql_upper or "OFFSET" in sql_upper or "FETCH NEXT" in sql_upper:
            return sql

        offset = (page - 1) * page_size

        # Check if sql ends with semicolon
        sql = sql.strip()
        if sql.endswith(";"):
            sql = sql[:-1]

        if db_type in [
            DatabaseType.MYSQL,
            DatabaseType.POSTGRESQL,
            DatabaseType.SQLITE,
            DatabaseType.MARIADB,
        ]:
            return f"{sql} LIMIT {page_size} OFFSET {offset}"
        elif db_type == DatabaseType.SQLSERVER:
            if "ORDER BY" in sql_upper:
                return f"{sql} OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY"
        elif db_type == DatabaseType.ORACLE:
            if "ORDER BY" in sql_upper:
                return f"{sql} OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY"

        return sql

    @staticmethod
    def execute_sql(user_id: str, request: SQLExecutionRequest) -> SQLExecutionResult:
        config_row = DatabaseToolService._get_config_with_password(
            request.db_config_id, user_id
        )
        if not config_row:
            return SQLExecutionResult(
                success=False,
                execution_time_ms=0,
                error_message="Configuration not found or access denied",
            )

        password, error = DatabaseToolService._decrypt_password(
            config_row["password_encrypted"], request.db_config_id
        )
        if error:
            return SQLExecutionResult(
                success=False,
                execution_time_ms=0,
                error_message=error,
            )

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name
            if request.database_name
            else config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
            "max_pool_size": config_row["max_pool_size"],
        }

        final_sql = request.sql

        # PostgreSQL schema 注入：当指定了 schema_name 时，自动为未指定 schema 的表名添加前缀
        if config_row["db_type"] == DatabaseType.POSTGRESQL and request.schema_name:
            from app.utils.sql_schema_injector import process_sql_with_schema_injection
            final_sql = process_sql_with_schema_injection(request.sql, request.schema_name)

        # 仅在单条 SELECT 语句时应用自动分页，避免影响多语句脚本
        statements = [
            s for s in sqlparse.split(final_sql)
            if is_executable_statement(s)
        ]

        if len(statements) == 1:
            if (
                request.page
                and request.page_size
                and request.page > 0
                and request.page_size > 0
            ):
                if statements[0].strip().upper().startswith("SELECT"):
                    final_sql = DatabaseToolService._apply_pagination(
                        statements[0],
                        request.page,
                        request.page_size,
                        config_row["db_type"],
                    )

        result = SQLExecutor.execute(
            request.db_config_id, config_dict, final_sql, request.params
        )

        # Save history
        DatabaseToolService._save_history(user_id, request, result)

        return result

    @staticmethod
    def _save_history(
        user_id: str, request: SQLExecutionRequest, result: SQLExecutionResult
    ):
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                history_id = str(uuid.uuid4())
                # 限制 result_data 大小，避免历史记录表膨胀和接口变慢（最大 100KB）
                MAX_RESULT_DATA_SIZE = 100 * 1024
                result_data_str = None
                result_size = 0
                if result.result_data:
                    result_data_str = json.dumps(result.result_data)
                    result_size = len(result_data_str)
                    if result_size > MAX_RESULT_DATA_SIZE:
                        truncated = result.result_data[:100] if isinstance(result.result_data, list) else result.result_data
                        result_data_str = json.dumps({
                            "__truncated__": True,
                            "original_size_bytes": result_size,
                            "message": "结果数据超过 100KB，仅保存摘要",
                            "preview": truncated,
                        })
                        result_size = len(result_data_str)

                cursor.execute(
                    """
                    INSERT INTO sql_execution_history (
                        id, user_id, db_config_id, sql_statement, sql_type,
                        execution_status, affected_rows, execution_time_ms,
                        error_message, result_data, result_size, created_at
                    ) VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, %s
                    )
                    """,
                    (
                        history_id,
                        user_id,
                        request.db_config_id,
                        request.sql,
                        result.sql_type,
                        "success" if result.success else "failed",
                        result.affected_rows,
                        int(result.execution_time_ms),
                        result.error_message,
                        result_data_str,
                        result_size,
                        datetime.now(),
                    ),
                )
                conn.commit()
        except Exception as e:
            logger.error(f"Failed to save execution history: {e}")
        finally:
            release_db_connection(conn)

    @staticmethod
    def get_history(
        user_id: str, limit: int = 50, offset: int = 0
    ) -> List[ExecutionHistory]:
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT h.*, c.alias as db_alias 
                    FROM sql_execution_history h
                    LEFT JOIN db_configs c ON h.db_config_id = c.id
                    WHERE h.user_id = %s 
                    ORDER BY h.created_at DESC 
                    LIMIT %s OFFSET %s
                    """,
                    (user_id, limit, offset),
                )
                rows = cursor.fetchall()

                history = []
                for row in rows:
                    history.append(
                        ExecutionHistory(
                            id=row["id"],
                            user_id=row["user_id"],
                            db_config_id=row["db_config_id"],
                            sql_statement=row["sql_statement"],
                            sql_type=row["sql_type"],
                            execution_status=row["execution_status"],
                            affected_rows=row["affected_rows"],
                            execution_time_ms=row["execution_time_ms"],
                            error_message=row["error_message"],
                            created_at=row["created_at"],
                            db_alias=row["db_alias"],
                        )
                    )
                return history
        finally:
            release_db_connection(conn)

    # --------------------------------------------------------------------------
    # Schema Browsing
    # --------------------------------------------------------------------------

    @staticmethod
    def get_databases_list(user_id: str, config_id: str, skip_cache: bool = False) -> List[str]:
        cache_key = f"databases:{config_id}"
        if not skip_cache:
            cached = _LIST_CACHE.get(cache_key)
            if cached is not None:
                return cached

        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        password, error = DatabaseToolService._decrypt_password(
            config_row["password_encrypted"], config_id
        )
        if error:
            raise ValueError(error)

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine = DBConnectionManager.get_engine(config_id, config_dict)

        db_type = config_row["db_type"]
        databases = []

        try:
            with engine.connect() as conn:
                if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                    result = conn.execute(text("SHOW DATABASES"))
                    databases = [row[0] for row in result]
                elif db_type == DatabaseType.POSTGRESQL:
                    # 如果配置中指定了 database_name，则返回该数据库下的 schema 列表
                    # 否则返回所有数据库的所有 schema（格式: "database:schema"）
                    target_db = config_row.get("database_name")
                    if target_db:
                        # 连接指定数据库，查询所有 schema
                        result = conn.execute(
                            text(
                                "SELECT schema_name FROM information_schema.schemata ORDER BY schema_name"
                            )
                        )
                        databases = [row[0] for row in result]
                    else:
                        # 只列库名：连 postgres 系统库查 pg_database，不再遍历各库查 schema。
                        # Schema 改由 get_schemas_list 在展开具体数据库时懒加载。
                        fallback_config = config_dict.copy()
                        fallback_config["database_name"] = "postgres"
                        temp_key = f"{config_id}:_temp_postgres_fallback"
                        fallback_engine = DBConnectionManager.get_engine(
                            temp_key, fallback_config
                        )
                        try:
                            with fallback_engine.connect() as fallback_conn:
                                db_result = fallback_conn.execute(
                                    text(
                                        "SELECT datname FROM pg_database WHERE datistemplate = false ORDER BY datname"
                                    )
                                )
                                databases = [row[0] for row in db_result]
                        finally:
                            fallback_engine.dispose()
                elif db_type == DatabaseType.SQLSERVER:
                    result = conn.execute(text("SELECT name FROM sys.databases"))
                    databases = [row[0] for row in result]
                elif db_type == DatabaseType.SQLITE:
                    databases = ["main"]
        except Exception as e:
            logger.error(f"Failed to list databases: {e}")
            # If we can't list databases, maybe we are restricted or connection failed
            # return empty list or re-raise
            raise e

        result = sorted(databases)
        _LIST_CACHE.set(cache_key, result)
        return result

    @staticmethod
    def get_schemas_list(
        user_id: str, config_id: str, database_name: Optional[str] = None,
        skip_cache: bool = False,
    ) -> List[str]:
        """获取指定数据库下的 schema 列表（PostgreSQL）"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        db_type = config_row["db_type"]
        if db_type != DatabaseType.POSTGRESQL:
            # 其他数据库不支持 schema 列表，返回空
            return []

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        # 确定目标数据库名：优先使用参数，其次使用配置中的 database_name
        target_db = database_name or config_row.get("database_name")
        if not target_db:
            raise ValueError("database_name is required for PostgreSQL schema listing")

        cache_key = f"schemas:{config_id}:{target_db}"
        if not skip_cache:
            cached = _LIST_CACHE.get(cache_key)
            if cached is not None:
                return cached

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": target_db,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
            "max_pool_size": config_row["max_pool_size"],
        }

        temp_key = f"{config_id}:_temp_schemas_{target_db}"
        engine = DBConnectionManager.get_engine(temp_key, config_dict)

        try:
            with engine.connect() as conn:
                result = conn.execute(
                    text("SELECT schema_name FROM information_schema.schemata ORDER BY schema_name")
                )
                schemas = [row[0] for row in result]
            _LIST_CACHE.set(cache_key, schemas)
            return schemas
        finally:
            engine.dispose()

    @staticmethod
    def get_all_schemas(
        user_id: str, config_id: str, skip_cache: bool = False
    ) -> Dict[str, List[str]]:
        """并行查询某 PG 连接下所有库的 schema，供搜索场景使用。

        返回 {database_name: [schema, ...]}。单库查询失败被隔离，记 warning 后跳过。
        """
        cache_key = f"all_schemas:{config_id}"
        if not skip_cache:
            cached = _LIST_CACHE.get(cache_key)
            if cached is not None:
                return cached

        db_names = DatabaseToolService.get_databases_list(
            user_id, config_id, skip_cache=skip_cache
        )

        result: Dict[str, List[str]] = {}

        def _fetch_one(db_name: str):
            return db_name, DatabaseToolService.get_schemas_list(
                user_id, config_id, db_name, skip_cache=skip_cache
            )

        # 限制并发，避免短时占用过多数据库连接
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(_fetch_one, db): db for db in db_names}
            for future in as_completed(futures):
                db = futures[future]
                try:
                    name, schemas = future.result()
                    result[name] = schemas
                except Exception as exc:
                    logger.warning(f"Failed to get schemas for {db}: {exc}")

        _LIST_CACHE.set(cache_key, result)
        return result

    @staticmethod
    def get_database_structure(
        user_id: str, config_id: str, database_name: str, schema_name: Optional[str] = None
    ) -> Dict[str, List[Dict[str, Any]]]:
        # 如果 database_name 包含 "database:schema" 格式，解析它
        actual_db_name = database_name
        actual_schema_name = schema_name
        if ":" in database_name:
            parts = database_name.split(":", 1)
            actual_db_name = parts[0]
            actual_schema_name = parts[1]

        cache_key = f"{config_id}:{actual_db_name}:{actual_schema_name}"

        # 1. 查后端缓存
        cached = _STRUCTURE_CACHE.get(cache_key)
        if cached:
            return cached

        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": actual_db_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        # Use a temporary config_id to avoid caching/conflict with main connection
        temp_config_id = f"{config_id}:{actual_db_name}"

        try:
            engine = DBConnectionManager.get_engine(temp_config_id, config_dict)
            db_type = config_row["db_type"]

            tables_data = []
            views_data = []

            if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                with engine.connect() as conn:
                    sql = text("""
                        SELECT TABLE_NAME, TABLE_COMMENT, 'table' AS obj_type
                        FROM information_schema.TABLES
                        WHERE TABLE_SCHEMA = :schema AND TABLE_TYPE = 'BASE TABLE'
                        UNION ALL
                        SELECT TABLE_NAME, NULL, 'view' AS obj_type
                        FROM information_schema.VIEWS
                        WHERE TABLE_SCHEMA = :schema
                        ORDER BY TABLE_NAME
                    """)
                    result = conn.execute(sql, {"schema": actual_db_name})
                    for row in result:
                        entry = {"name": row[0], "comment": row[1]}
                        if row[2] == 'table':
                            tables_data.append(entry)
                        else:
                            views_data.append(entry)

            elif db_type == DatabaseType.POSTGRESQL:
                pg_schema = actual_schema_name or "public"
                with engine.connect() as conn:
                    # 查询表（包含注释）
                    sql_tables = text("""
                        SELECT c.relname AS table_name,
                               obj_description(c.oid, 'pg_class') AS table_comment
                        FROM pg_class c
                        JOIN pg_namespace n ON n.oid = c.relnamespace
                        WHERE c.relkind = 'r'  -- regular table
                        AND n.nspname = :schema
                        ORDER BY c.relname
                    """)
                    result = conn.execute(sql_tables, {"schema": pg_schema})
                    for row in result:
                        tables_data.append({"name": row[0], "comment": row[1]})

                    # 查询视图
                    sql_views = text("""
                        SELECT c.relname AS view_name
                        FROM pg_class c
                        JOIN pg_namespace n ON n.oid = c.relnamespace
                        WHERE c.relkind = 'v'  -- view
                        AND n.nspname = :schema
                        ORDER BY c.relname
                    """)
                    result = conn.execute(sql_views, {"schema": pg_schema})
                    for row in result:
                        views_data.append({"name": row[0], "comment": None})

            else:
                # 其他数据库类型保持原有逻辑
                inspector = inspect(engine)
                for name in inspector.get_table_names():
                    try:
                        comment = inspector.get_table_comment(name).get("text")
                    except:
                        comment = None
                    tables_data.append({"name": name, "comment": comment})

                for name in inspector.get_view_names():
                    views_data.append({"name": name, "comment": None})

            result = {"tables": tables_data, "views": views_data}
            _STRUCTURE_CACHE.set(cache_key, result)
            return result
        except Exception as e:
            logger.error(f"Failed to get database structure: {e}")
            raise e

    @staticmethod
    def get_tables(user_id: str, config_id: str) -> List[str]:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine = DBConnectionManager.get_engine(config_id, config_dict)
        inspector = inspect(engine)
        return inspector.get_table_names()

    @staticmethod
    def get_table_schema(
        user_id: str,
        config_id: str,
        table_name: str,
        database_name: Optional[str] = None,
        schema_name: Optional[str] = None,
    ) -> TableSchema:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        target_db = database_name if database_name else config_row["database_name"]

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": target_db,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        # Use a unique key if database_name is provided to avoid using cached engine of default DB
        engine_key = f"{config_id}:{target_db}" if database_name else config_id

        # DBConnectionManager now handles key generation based on database_name
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        inspector = inspect(engine)

        # PostgreSQL: 使用 schema 关键字参数
        is_postgresql = config_row["db_type"] == DatabaseType.POSTGRESQL
        schema_kw = {"schema": schema_name} if (is_postgresql and schema_name) else {}

        # Get table comment
        table_comment = None
        try:
            table_comment = inspector.get_table_comment(table_name, **schema_kw).get("text")
        except:
            pass

        columns = []
        for col in inspector.get_columns(table_name, **schema_kw):
            # Convert SQLAlchemy type to string
            col["type"] = str(col["type"])
            # Ensure comment is present
            if "comment" not in col:
                col["comment"] = None
            columns.append(col)

        pk_constraint = inspector.get_pk_constraint(table_name, **schema_kw)
        primary_key = pk_constraint.get("constrained_columns", [])

        indexes = inspector.get_indexes(table_name, **schema_kw)
        foreign_keys = inspector.get_foreign_keys(table_name, **schema_kw)

        return TableSchema(
            table_name=table_name,
            comment=table_comment,
            columns=columns,
            primary_key=primary_key,
            indexes=indexes,
            foreign_keys=foreign_keys,
        )

    @staticmethod
    def query_table_data(
        user_id: str,
        config_id: str,
        table_name: str,
        database_name: Optional[str] = None,
        schema_name: Optional[str] = None,
        where_clause: Optional[str] = None,
        order_by_clause: Optional[str] = None,
        page: int = 1,
        page_size: int = 20,
    ) -> SQLExecutionResult:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            return SQLExecutionResult(
                success=False,
                execution_time_ms=0,
                error_message="Configuration not found",
            )

        password, error = DatabaseToolService._decrypt_password(
            config_row["password_encrypted"], config_id
        )
        if error:
            return SQLExecutionResult(
                success=False,
                execution_time_ms=0,
                error_message=error,
            )

        # If database_name is provided (e.g. for multi-db connections), use it
        # Otherwise use the default from config
        target_db = database_name if database_name else config_row["database_name"]

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": target_db,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
            "max_pool_size": config_row["max_pool_size"],
        }

        # We need a unique engine key if we are switching databases dynamically
        engine_key = f"{config_id}:{target_db}" if database_name else config_id

        # Construct SQL
        # Note: This is a direct SQL construction.
        # Since this is a DB tool for developers, we allow raw WHERE/ORDER BY clauses.
        # But we should be careful about injection if this was a public app.

        # Basic validation to prevent simple injection of "; DROP TABLE" style attacks if possible,
        # though user has SQL execution rights anyway via the main tool.

        db_type = config_row["db_type"]

        # Schema-qualified table name for PostgreSQL
        if db_type == DatabaseType.POSTGRESQL and schema_name:
            qualified_table = f'"{schema_name}"."{table_name}"'
        else:
            qualified_table = table_name

        sql = f"SELECT * FROM {qualified_table}"
        count_sql = f"SELECT COUNT(*) FROM {qualified_table}"

        if where_clause and where_clause.strip():
            sql += f" WHERE {where_clause}"
            count_sql += f" WHERE {where_clause}"

        if order_by_clause and order_by_clause.strip():
            sql += f" ORDER BY {order_by_clause}"

        # Pagination
        offset = (page - 1) * page_size

        # Use SQLExecutor logic but adapted for pagination
        # Or just construct the limit/offset string based on DB type
        # SQLAlchemy handles this if we use select().limit().offset(), but we are using raw strings for where clause.
        # Most DBs support LIMIT/OFFSET or similar.

        db_type = config_row["db_type"]

        if db_type == DatabaseType.SQLSERVER:
            # SQL Server 2012+ supports OFFSET/FETCH
            # It requires ORDER BY to use OFFSET/FETCH
            if not order_by_clause or not order_by_clause.strip():
                # Need default order by for pagination
                sql += " ORDER BY (SELECT NULL)"
            sql += f" OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY"
        elif db_type == DatabaseType.ORACLE:
            # Oracle 12c+
            sql += f" OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY"
        else:
            # MySQL, PostgreSQL, SQLite
            sql += f" LIMIT {page_size} OFFSET {offset}"

        return SQLExecutor.execute(engine_key, config_dict, sql)

    # --------------------------------------------------------------------------
    # Database Administration (DDL)
    # --------------------------------------------------------------------------

    @staticmethod
    def create_database_instance(
        user_id: str, config_id: str, database_name: str, charset: str = "utf8mb4"
    ) -> bool:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        # Connect to the default database of the connection to execute CREATE DATABASE
        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": config_row["database_name"],  # Connect to default DB
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        db_type = config_row["db_type"]
        sql = ""

        # Simple sanitization for database name (alphanumeric + underscore)
        # In a real tool, we should use proper quoting/escaping
        if not database_name.replace("_", "").isalnum():
            # Fallback or strict check. For now, let's assume valid input or basic quoting.
            # But better to quote it.
            pass

        # Quoting helper
        def quote_ident(name):
            return f'"{name}"' if db_type == DatabaseType.POSTGRESQL else f"`{name}`"

        quoted_name = quote_ident(database_name)

        if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
            sql = f"CREATE DATABASE {quoted_name} CHARACTER SET {charset}"
        elif db_type == DatabaseType.POSTGRESQL:
            sql = f"CREATE DATABASE {quoted_name}"
            # Postgres cannot execute CREATE DATABASE inside a transaction block
            # We need to handle this. SQLAlchemy's execute usually auto-commits or we need execution_options(isolation_level="AUTOCOMMIT")
        elif db_type == DatabaseType.SQLSERVER:
            sql = f"CREATE DATABASE {quoted_name}"
        elif db_type == DatabaseType.SQLITE:
            # SQLite doesn't really "CREATE DATABASE" in the same way (it's a file)
            # We could ATTACH, but usually we just create a file.
            # For now, skip or return False
            return False

        engine = DBConnectionManager.get_engine(config_id, config_dict)

        try:
            # For DDL that can't run in transaction (Postgres CREATE DB), we need autocommit
            with engine.connect().execution_options(
                isolation_level="AUTOCOMMIT"
            ) as conn:
                conn.execute(text(sql))
            return True
        except Exception as e:
            logger.error(f"Failed to create database: {e}")
            raise e

    @staticmethod
    def drop_database_instance(
        user_id: str, config_id: str, database_name: str
    ) -> bool:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": config_row["database_name"],  # Connect to default DB
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        db_type = config_row["db_type"]

        def quote_ident(name):
            return f'"{name}"' if db_type == DatabaseType.POSTGRESQL else f"`{name}`"

        quoted_name = quote_ident(database_name)
        sql = f"DROP DATABASE {quoted_name}"

        engine = DBConnectionManager.get_engine(config_id, config_dict)

        try:
            with engine.connect().execution_options(
                isolation_level="AUTOCOMMIT"
            ) as conn:
                conn.execute(text(sql))
            return True
        except Exception as e:
            logger.error(f"Failed to drop database: {e}")
            raise e

    @staticmethod
    def drop_table_instance(
        user_id: str, config_id: str, database_name: str, table_name: str
    ) -> bool:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": database_name,  # Connect to specific DB
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        db_type = config_row["db_type"]

        def quote_ident(name):
            return f'"{name}"' if db_type == DatabaseType.POSTGRESQL else f"`{name}`"

        quoted_table = quote_ident(table_name)
        sql = f"DROP TABLE {quoted_table}"

        # Unique key for engine because we might switch databases
        engine_key = f"{config_id}:{database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)

        try:
            with engine.connect().execution_options(
                isolation_level="AUTOCOMMIT"
            ) as conn:
                conn.execute(text(sql))
            return True
        except Exception as e:
            logger.error(f"Failed to drop table: {e}")
            raise e

    @staticmethod
    def truncate_table_instance(
        user_id: str, config_id: str, database_name: str, table_name: str
    ) -> bool:
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        db_type = config_row["db_type"]

        def quote_ident(name):
            return f'"{name}"' if db_type == DatabaseType.POSTGRESQL else f"`{name}`"

        quoted_table = quote_ident(table_name)

        if db_type == DatabaseType.SQLITE:
            sql = f"DELETE FROM {quoted_table}"
        else:
            sql = f"TRUNCATE TABLE {quoted_table}"

        engine_key = f"{config_id}:{database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)

        # 使用 engine.begin() 模式 + 自动重连机制
        # 解决远程数据库连接断开时 "server closed the connection unexpectedly" 错误
        last_error = None
        for attempt in range(2):  # 最多重试 1 次
            try:
                with engine.begin() as conn:
                    conn.execute(text(sql))
                return True
            except Exception as e:
                last_error = e
                error_msg = str(e).lower()
                # 连接断开错误，丢弃失效引擎并重试
                if any(keyword in error_msg for keyword in [
                    "server closed the connection",
                    "connection closed",
                    "connection refused",
                    "lost connection",
                    "broken pipe",
                    "connection reset",
                    "ssl connection has been closed",
                ]):
                    logger.warning(
                        f"Truncate table 连接已断开 (尝试 {attempt + 1}/2): {e}"
                    )
                    if attempt == 0:
                        # 关闭失效引擎，下次会重新创建
                        DBConnectionManager.close_engine(engine_key)
                        engine = DBConnectionManager.get_engine(engine_key, config_dict)
                        continue
                # 非连接错误，直接抛出
                logger.error(f"Failed to truncate table: {e}")
                raise

        # 重试后仍失败
        logger.error(f"Failed to truncate table after retry: {last_error}")
        raise last_error

    # --------------------------------------------------------------------------
    # Search
    # --------------------------------------------------------------------------

    @staticmethod
    def search_tables(
        user_id: str, config_id: str, keyword: str
    ) -> List[Dict[str, str]]:
        """
        Search for tables and views matching the keyword across all databases (if supported).
        Returns a list of { "database": "db_name", "table": "table_name", "type": "table|view" }
        """
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        db_type = config_row["db_type"]
        results = []

        try:
            # MySQL/MariaDB: Use information_schema for global search
            if db_type == DatabaseType.MYSQL or db_type == DatabaseType.MARIADB:
                engine = DBConnectionManager.get_engine(config_id, config_dict)
                with engine.connect() as conn:
                    # Search tables and views
                    sql = text("""
                        SELECT TABLE_SCHEMA, TABLE_NAME, TABLE_TYPE 
                        FROM information_schema.TABLES 
                        WHERE TABLE_NAME LIKE :keyword
                        AND TABLE_SCHEMA NOT IN ('information_schema', 'mysql', 'performance_schema', 'sys')
                        ORDER BY TABLE_SCHEMA, TABLE_NAME
                        LIMIT 100
                    """)
                    rows = conn.execute(sql, {"keyword": f"%{keyword}%"}).fetchall()
                    for row in rows:
                        results.append(
                            {
                                "database": row[0],
                                "table": row[1],
                                "type": "view"
                                if "VIEW" in str(row[2]).upper()
                                else "table",
                            }
                        )

            # PostgreSQL: Search in current DB, return schema info
            elif db_type == DatabaseType.POSTGRESQL:
                engine = DBConnectionManager.get_engine(config_id, config_dict)
                with engine.connect() as conn:
                    sql = text("""
                        SELECT table_schema, table_name, table_type
                        FROM information_schema.tables
                        WHERE table_name ILIKE :keyword
                        AND table_schema NOT IN ('information_schema', 'pg_catalog')
                        LIMIT 100
                    """)
                    rows = conn.execute(sql, {"keyword": f"%{keyword}%"}).fetchall()
                    current_db = config_row["database_name"]
                    for row in rows:
                        results.append(
                            {
                                "database": f"{current_db}:{row[0]}",  # Include schema info
                                "table": row[1],
                                "type": "view"
                                if "VIEW" in str(row[2]).upper()
                                else "table",
                            }
                        )

            # SQL Server: Search in current DB
            elif db_type == DatabaseType.SQLSERVER:
                engine = DBConnectionManager.get_engine(config_id, config_dict)
                with engine.connect() as conn:
                    sql = text("""
                        SELECT TABLE_CATALOG, TABLE_NAME, TABLE_TYPE
                        FROM information_schema.TABLES
                        WHERE TABLE_NAME LIKE :keyword
                        LIMIT 100
                    """)
                    rows = conn.execute(sql, {"keyword": f"%{keyword}%"}).fetchall()
                    for row in rows:
                        results.append(
                            {
                                "database": row[0],
                                "table": row[1],
                                "type": "view"
                                if "VIEW" in str(row[2]).upper()
                                else "table",
                            }
                        )

            # SQLite
            elif db_type == DatabaseType.SQLITE:
                engine = DBConnectionManager.get_engine(config_id, config_dict)
                with engine.connect() as conn:
                    sql = text("""
                        SELECT name, type FROM sqlite_master 
                        WHERE name LIKE :keyword AND type IN ('table', 'view')
                        LIMIT 100
                    """)
                    rows = conn.execute(sql, {"keyword": f"%{keyword}%"}).fetchall()
                    for row in rows:
                        results.append(
                            {"database": "main", "table": row[0], "type": row[1]}
                        )

        except Exception as e:
            logger.error(f"Search failed: {e}")
            # Don't fail the whole request, just return empty or what we found

        return results

    # ============ 数据导入导出功能 ============

    @staticmethod
    def export_data(
        user_id: str, config_id: str, request: ExportDataRequest
    ) -> ExportDataResponse:
        """导出数据为指定格式"""
        logger.info(f"Export request: config_id={config_id}, format={request.format}, sql={request.sql[:200]}")
        config_row = DatabaseToolService._get_config_with_password(
            config_id, user_id
        )
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name
            if request.database_name
            else config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = (
            f"{config_id}:{request.database_name}"
            if request.database_name
            else config_id
        )
        engine = DBConnectionManager.get_engine(engine_key, config_dict)

        import csv
        import io
        import json

        try:
            with engine.connect() as conn:
                result = conn.execute(text(request.sql))
                rows = result.fetchall()
                columns = list(result.keys())

                if request.format == ExportFormat.CSV:
                    output = io.StringIO()
                    writer = csv.writer(output)
                    writer.writerow(columns)
                    for row in rows:
                        writer.writerow(row)
                    content = output.getvalue()
                    file_name = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    return ExportDataResponse(
                        file_name=file_name,
                        file_size=len(content.encode("utf-8")),
                        content=content,
                        row_count=len(rows),
                    )

                elif request.format == ExportFormat.JSON:
                    data = [dict(zip(columns, row)) for row in rows]
                    content = json.dumps(
                        data, indent=2, ensure_ascii=False, default=str
                    )
                    file_name = (
                        f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                    )
                    return ExportDataResponse(
                        file_name=file_name,
                        file_size=len(content.encode("utf-8")),
                        content=content,
                        row_count=len(rows),
                    )

                elif request.format == ExportFormat.EXCEL:
                    # 需要 xlsx 库
                    try:
                        from openpyxl import Workbook

                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Export"

                        # 添加列头
                        ws.append(columns)

                        # 添加数据
                        for row in rows:
                            ws.append(
                                [
                                    str(cell) if cell is not None else None
                                    for cell in row
                                ]
                            )

                        # 保存到字节
                        import io

                        output = io.BytesIO()
                        wb.save(output)
                        output.seek(0)

                        import base64

                        content = base64.b64encode(output.getvalue()).decode("utf-8")
                        file_name = (
                            f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        )
                        return ExportDataResponse(
                            file_name=file_name,
                            file_size=len(content),
                            content=content,
                            row_count=len(rows),
                        )
                    except ImportError:
                        raise ValueError(
                            "openpyxl library not installed. Please install it first."
                        )

                elif request.format == ExportFormat.SQL:
                    # 生成 INSERT 语句
                    import re
                    # 从 SQL 中提取表名（支持大小写）
                    table_match = re.search(
                        r'FROM\s+([`"\']?\w+[`"\']?(?:\.[`"\']?\w+[`"\']?)?)',
                        request.sql, re.IGNORECASE
                    )
                    table_name = table_match.group(1) if table_match else "exported_data"

                    statements = []
                    for row in rows:
                        values = []
                        for val in row:
                            if val is None:
                                values.append("NULL")
                            elif isinstance(val, str):
                                values.append(f"'{val.replace(chr(39), chr(39)+chr(39))}'")
                            elif isinstance(val, (datetime,)):
                                values.append(f"'{val}'")
                            else:
                                values.append(str(val))
                        col_list = ", ".join(columns)
                        statements.append(
                            f"INSERT INTO {table_name} ({col_list}) VALUES ({', '.join(values)})"
                        )

                    content = ";\n".join(statements) + ";" if statements else ""
                    file_name = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
                    return ExportDataResponse(
                        file_name=file_name,
                        file_size=len(content.encode("utf-8")),
                        content=content,
                        row_count=len(rows),
                    )

                else:
                    raise ValueError(f"Unsupported format: {request.format}")

        except Exception as e:
            logger.error(f"Export failed: {e}")
            raise e

    @staticmethod
    def import_data(
        user_id: str, config_id: str, request: ImportDataRequest
    ) -> ImportDataResponse:
        """从文件导入数据"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name
            if request.database_name
            else config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = (
            f"{config_id}:{request.database_name}"
            if request.database_name
            else config_id
        )
        engine = DBConnectionManager.get_engine(engine_key, config_dict)

        import csv
        import io
        import json

        errors = []
        imported_rows = 0
        skipped_rows = 0

        try:
            if request.format == ExportFormat.CSV:
                output = io.StringIO(request.content)
                reader = csv.reader(output)
                columns = next(reader)

                with engine.connect() as conn:
                    if request.overwrite:
                        conn.execute(text(f"TRUNCATE TABLE {request.table_name}"))
                        conn.commit()

                    for row_num, row in enumerate(reader, start=2):
                        try:
                            values = ", ".join(["%s"] * len(row))
                            sql = f"INSERT INTO {request.table_name} ({', '.join(columns)}) VALUES ({values})"
                            conn.execute(text(sql), tuple(row))
                            imported_rows += 1
                        except Exception as e:
                            errors.append({"row": row_num, "error": str(e)})
                            skipped_rows += 1

                    conn.commit()

            elif request.format == ExportFormat.JSON:
                data = json.loads(request.content)
                if not data:
                    raise ValueError("Empty JSON data")

                with engine.connect() as conn:
                    if request.overwrite:
                        conn.execute(text(f"TRUNCATE TABLE {request.table_name}"))
                        conn.commit()

                    for item in data:
                        try:
                            columns = list(item.keys())
                            values = list(item.values())
                            values_str = ", ".join(["%s"] * len(values))
                            sql = f"INSERT INTO {request.table_name} ({', '.join(columns)}) VALUES ({values_str})"
                            conn.execute(text(sql), tuple(values))
                            imported_rows += 1
                        except Exception as e:
                            errors.append({"item": str(item)[:50], "error": str(e)})
                            skipped_rows += 1

                    conn.commit()

            elif request.format == ExportFormat.EXCEL:
                try:
                    from openpyxl import load_workbook
                    import base64

                    file_data = base64.b64decode(request.content)
                    wb = load_workbook(filename=io.BytesIO(file_data))
                    ws = wb.active

                    columns = [cell.value for cell in ws[1]]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(row)

                    with engine.connect() as conn:
                        if request.overwrite:
                            conn.execute(text(f"TRUNCATE TABLE {request.table_name}"))
                            conn.commit()

                        for row_num, row in enumerate(rows, start=2):
                            try:
                                values = [
                                    str(cell) if cell is not None else None
                                    for cell in row
                                ]
                                values_str = ", ".join(["%s"] * len(values))
                                sql = f"INSERT INTO {request.table_name} ({', '.join(columns)}) VALUES ({values_str})"
                                conn.execute(text(sql), tuple(values))
                                imported_rows += 1
                            except Exception as e:
                                errors.append({"row": row_num, "error": str(e)})
                                skipped_rows += 1

                        conn.commit()

                except ImportError:
                    raise ValueError("openpyxl library not installed.")

            return ImportDataResponse(
                success=len(errors) == 0,
                imported_rows=imported_rows,
                skipped_rows=skipped_rows,
                errors=errors,
            )

        except Exception as e:
            logger.error(f"Import failed: {e}")
            raise e

    # ============ 执行计划分析 ============

    @staticmethod
    def explain_plan(
        user_id: str, config_id: str, request: ExplainPlanRequest
    ) -> ExplainPlanResponse:
        """分析 SQL 执行计划"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name
            if request.database_name
            else config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = (
            f"{config_id}:{request.database_name}"
            if request.database_name
            else config_id
        )
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        db_type = config_row["db_type"]

        start_time = time.time()

        try:
            with engine.connect() as conn:
                if db_type in [DatabaseType.MYSQL, DatabaseType.MARIADB]:
                    result = conn.execute(text(f"EXPLAIN {request.sql}"))
                    rows = result.fetchall()
                    columns = result.keys()

                    plan = []
                    for row in rows:
                        plan.append(
                            ExplainPlanStep(
                                id=row[columns.index("id")] if "id" in columns else 0,
                                select_type=row[columns.index("select_type")]
                                if "select_type" in columns
                                else None,
                                table=row[columns.index("table")]
                                if "table" in columns
                                else None,
                                type=row[columns.index("type")]
                                if "type" in columns
                                else None,
                                possible_keys=row[columns.index("possible_keys")]
                                if "possible_keys" in columns
                                else None,
                                key=row[columns.index("key")]
                                if "key" in columns
                                else None,
                                key_len=row[columns.index("key_len")]
                                if "key_len" in columns
                                else None,
                                ref=row[columns.index("ref")]
                                if "ref" in columns
                                else None,
                                rows=row[columns.index("rows")]
                                if "rows" in columns
                                else None,
                                filtered=row[columns.index("Filtered")]
                                if "Filtered" in columns
                                else None,
                                extra=row[columns.index("Extra")]
                                if "Extra" in columns
                                else None,
                            )
                        )

                elif db_type == DatabaseType.POSTGRESQL:
                    result = conn.execute(text(f"EXPLAIN (FORMAT JSON) {request.sql}"))
                    row = result.fetchone()
                    # PostgreSQL 返回 JSON 格式的执行计划
                    plan_data = row[0] if row else []
                    plan = []
                    # 简化处理，实际需要解析 JSON
                    for step in plan_data:
                        plan.append(
                            ExplainPlanStep(
                                id=0,
                                select_type=step.get("Node Type"),
                                table=step.get("Relation Name"),
                                type=None,
                                rows=step.get("Plan Rows"),
                                extra=str(step),
                            )
                        )

                elif db_type == DatabaseType.SQLSERVER:
                    result = conn.execute(text(f"SET SHOWPLAN_TEXT ON; {request.sql}"))
                    rows = result.fetchall()
                    plan = [
                        ExplainPlanStep(
                            id=i,
                            select_type=None,
                            table=str(row[0]) if row else None,
                            type=None,
                        )
                        for i, row in enumerate(rows)
                    ]

                else:
                    raise ValueError(f"EXPLAIN not supported for {db_type}")

            execution_time = (time.time() - start_time) * 1000

            # 生成简单的分析建议
            analysis = DatabaseToolService._analyze_plan(plan)

            return ExplainPlanResponse(
                success=True,
                plan=plan,
                analysis=analysis,
                execution_time_ms=execution_time,
            )

        except Exception as e:
            logger.error(f"Explain plan failed: {e}")
            return ExplainPlanResponse(
                success=False, plan=[], analysis=None, execution_time_ms=0
            )

    @staticmethod
    def _analyze_plan(plan: List[ExplainPlanStep]) -> str:
        """分析执行计划并给出建议"""
        suggestions = []

        for step in plan:
            if step.rows and step.rows > 10000:
                suggestions.append(
                    f"表 {step.table} 扫描行数较多 ({step.rows})，考虑添加索引"
                )

            if step.type and step.type in ["ALL", "index"]:
                suggestions.append(f"表 {step.table} 进行全表扫描，建议检查索引")

            if step.key is None and step.table:
                suggestions.append(f"表 {step.table} 未使用索引")

            if step.extra and "Using temporary" in step.extra:
                suggestions.append("查询使用了临时表，考虑优化查询语句")

            if step.extra and "Using filesort" in step.extra:
                suggestions.append("查询使用了文件排序，考虑在 ORDER BY 字段添加索引")

        return "\\n".join(suggestions) if suggestions else "执行计划良好，无需优化"

    # ============ 表数据预览 ============

    @staticmethod
    def table_preview(
        user_id: str, config_id: str, request: TablePreviewRequest
    ) -> TablePreviewResponse:
        """预览表数据"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{request.database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        db_type = config_row["db_type"]

        try:
            with engine.connect() as conn:
                # 获取总行数
                count_result = conn.execute(
                    text(f"SELECT COUNT(*) FROM {request.table_name}")
                )
                total_count = count_result.scalar()

                # 构建查询
                sql = f"SELECT * FROM {request.table_name}"

                if request.order_by:
                    sql += f" ORDER BY {request.order_by}"

                # 分页
                offset = (request.page - 1) * request.page_size

                if db_type in [
                    DatabaseType.MYSQL,
                    DatabaseType.MARIADB,
                    DatabaseType.POSTGRESQL,
                    DatabaseType.SQLITE,
                ]:
                    sql += f" LIMIT {request.page_size} OFFSET {offset}"
                elif db_type == DatabaseType.SQLSERVER:
                    sql += f" OFFSET {offset} ROWS FETCH NEXT {request.page_size} ROWS ONLY"

                result = conn.execute(text(sql))
                rows = result.fetchall()
                columns = result.keys()

                data = [dict(zip(columns, row)) for row in rows]

                return TablePreviewResponse(
                    columns=list(columns),
                    rows=data,
                    total_count=total_count,
                    page=request.page,
                    page_size=request.page_size,
                    has_more=(request.page * request.page_size) < total_count,
                )

        except Exception as e:
            logger.error(f"Table preview failed: {e}")
            raise e

    # ============ SQL 自动补全 ============

    @staticmethod
    def auto_complete(
        user_id: str, config_id: str, request: AutoCompleteRequest
    ) -> AutoCompleteResponse:
        """SQL 自动补全"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name
            if request.database_name
            else config_row["database_name"],
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = (
            f"{config_id}:{request.database_name}"
            if request.database_name
            else config_id
        )
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        db_type = config_row["db_type"]

        suggestions = []

        # SQL 关键字补全
        keywords = [
            "SELECT",
            "FROM",
            "WHERE",
            "JOIN",
            "LEFT JOIN",
            "RIGHT JOIN",
            "INNER JOIN",
            "ORDER BY",
            "GROUP BY",
            "HAVING",
            "LIMIT",
            "OFFSET",
            "INSERT INTO",
            "UPDATE",
            "DELETE FROM",
            "CREATE TABLE",
            "ALTER TABLE",
            "DROP TABLE",
        ]

        query_upper = request.query.upper()
        for kw in keywords:
            if kw.startswith(query_upper.split()[-1] if query_upper.split() else ""):
                suggestions.append(
                    AutoCompleteItem(label=kw, kind="keyword", insert_text=kw)
                )

        try:
            with engine.connect() as conn:
                # 获取表名
                if db_type in [DatabaseType.MYSQL, DatabaseType.MARIADB]:
                    result = conn.execute(
                        text("""
                        SELECT TABLE_NAME, TABLE_COMMENT
                        FROM information_schema.TABLES
                        WHERE TABLE_SCHEMA = DATABASE()
                        LIMIT 100
                    """)
                    )
                elif db_type == DatabaseType.POSTGRESQL:
                    schema_filter = request.schema_name or "public"
                    result = conn.execute(
                        text("""
                        SELECT tablename, ''
                        FROM pg_tables
                        WHERE schemaname = :schema
                        LIMIT 100
                    """).bindparams(schema=schema_filter)
                    )
                elif db_type == DatabaseType.SQLITE:
                    result = conn.execute(
                        text("""
                        SELECT name, ''
                        FROM sqlite_master
                        WHERE type='table'
                        LIMIT 100
                    """)
                    )
                else:
                    result = []

                for row in result:
                    suggestions.append(
                        AutoCompleteItem(
                            label=row[0],
                            kind="table",
                            detail=row[1] if len(row) > 1 else None,
                            insert_text=row[0],
                        )
                    )

                    # 获取列名（当输入包含表名时）
                    if row[0] in request.query:
                        try:
                            from sqlalchemy import inspect

                            inspector = inspect(engine)
                            for col in inspector.get_columns(row[0]):
                                suggestions.append(
                                    AutoCompleteItem(
                                        label=col["name"],
                                        kind="column",
                                        detail=str(col["type"]),
                                        insert_text=col["name"],
                                    )
                                )
                        except:
                            pass

        except Exception as e:
            logger.error(f"Get auto complete failed: {e}")

        return AutoCompleteResponse(suggestions=suggestions[:50])  # 限制返回数量

    # ============ 备份恢复功能 ============

    @staticmethod
    def backup_database(
        user_id: str, config_id: str, request: BackupDatabaseRequest
    ) -> BackupDatabaseResponse:
        """备份数据库 — 支持全数据库类型和多种备份模式"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{request.database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        db_type = config_row["db_type"]

        generator = get_generator(db_type)

        try:
            with engine.connect() as conn:
                inspector = inspect(engine)
                tables = (
                    request.tables if request.tables else inspector.get_table_names()
                )

                sql_statements = []
                mode = request.backup_mode.value if hasattr(request.backup_mode, 'value') else str(request.backup_mode)

                for table in tables:
                    if request.include_drop:
                        sql_statements.append(f"DROP TABLE IF EXISTS `{table}`;")

                    # 结构
                    if mode in ("structure_and_data", "structure_only"):
                        ddl = generator.get_create_table_ddl(conn, table, request.database_name)
                        if request.include_if_not_exists and "CREATE TABLE" in ddl:
                            ddl = ddl.replace("CREATE TABLE", "CREATE TABLE IF NOT EXISTS", 1)
                        sql_statements.append(ddl + ";")

                    # 数据
                    if mode in ("structure_and_data", "data_only"):
                        inserts = generator.get_insert_statements(conn, table, request.database_name)
                        sql_statements.extend(inserts)

                content = "\n\n".join(sql_statements)

                # 保存到服务端存储
                record = backup_storage.save_backup(
                    user_id=user_id,
                    config_id=config_id,
                    database_name=request.database_name,
                    backup_mode=mode,
                    tables=tables,
                    sql_content=content,
                )

                return BackupDatabaseResponse(
                    backup_id=record.id,
                    file_name=record.file_name,
                    file_size=record.file_size,
                    download_url=f"/api/database-tool/backups/{record.id}/download",
                    created_at=datetime.fromisoformat(record.created_at),
                    tables_count=record.tables_count,
                    backup_mode=record.backup_mode,
                    status="success",
                )

        except Exception as e:
            logger.error(f"Backup failed: {e}")
            raise e

    @staticmethod
    def restore_database(
        user_id: str, config_id: str, request: RestoreDatabaseRequest
    ) -> RestoreDatabaseResponse:
        """恢复数据库"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": request.target_database,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{request.target_database}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)

        start_time = time.time()
        errors = []
        restored_tables = []
        total_rows = 0

        try:
            with engine.connect() as conn:
                # 简单分割 SQL 语句
                statements = request.backup_file_content.split(";")

                for stmt in statements:
                    stmt = stmt.strip()
                    if stmt:
                        try:
                            conn.execute(text(stmt))
                            if stmt.upper().startswith("INSERT"):
                                total_rows += 1
                            elif stmt.upper().startswith("CREATE TABLE"):
                                # 提取表名
                                parts = stmt.split()
                                if len(parts) > 2:
                                    table_name = parts[2].strip('`"')
                                    restored_tables.append(table_name)
                        except Exception as e:
                            errors.append(str(e))

                conn.commit()

            execution_time = (time.time() - start_time) * 1000

            return RestoreDatabaseResponse(
                success=len(errors) == 0,
                restored_tables=list(set(restored_tables)),
                total_rows=total_rows,
                execution_time_ms=execution_time,
                errors=errors,
            )

        except Exception as e:
            logger.error(f"Restore failed: {e}")
            return RestoreDatabaseResponse(
                success=False,
                restored_tables=[],
                total_rows=0,
                execution_time_ms=0,
                errors=[str(e)],
            )

    @staticmethod
    def batch_delete_rows(
        user_id: str, config_id: str, table_name: str, request: "BatchDeleteRequest"
    ) -> "BatchDeleteResult":
        from app.models.database_tool_models import (
            BatchDeleteRequest,
            BatchDeleteResult,
        )

        start_time = datetime.now()

        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            return BatchDeleteResult(
                success=False,
                deleted_count=0,
                failed_count=0,
                error_message="Configuration not found",
                execution_time_ms=0,
            )

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            return BatchDeleteResult(
                success=False,
                deleted_count=0,
                failed_count=0,
                error_message="Failed to decrypt password",
                execution_time_ms=0,
            )

        target_db = (
            request.database_name
            if request.database_name
            else config_row["database_name"]
        )

        target_schema = getattr(request, 'schema_name', None)

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": target_db,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
            "max_pool_size": config_row["max_pool_size"],
        }

        engine_key = f"{config_id}:{target_db}" if request.database_name else config_id
        db_type = config_row["db_type"]

        def quote_ident(name: str) -> str:
            if db_type == DatabaseType.SQLSERVER:
                return f"[{name}]"
            if db_type == DatabaseType.POSTGRESQL:
                return f'"{name}"'
            return f"`{name}`"

        def escape_value(val: Any) -> str:
            if val is None:
                return "NULL"
            if isinstance(val, bool):
                if db_type == DatabaseType.POSTGRESQL:
                    return "TRUE" if val else "FALSE"
                return "1" if val else "0"
            if isinstance(val, (int, float)):
                return str(val)
            if isinstance(val, str) and val.isdigit():
                return val
            escaped = str(val).replace("'", "''")
            return f"'{escaped}'"

        if target_schema and db_type == DatabaseType.POSTGRESQL:
            table_quoted = f'{quote_ident(target_schema)}.{quote_ident(table_name)}'
        else:
            table_quoted = quote_ident(table_name)

        if len(request.primary_keys) == 1:
            pk_col = quote_ident(request.primary_keys[0])
            values = [
                escape_value(row[request.primary_keys[0]]) for row in request.key_values
            ]
            where_clause = f"{pk_col} IN ({', '.join(values)})"
        else:
            pk_cols = ", ".join(quote_ident(k) for k in request.primary_keys)
            tuples = []
            for row in request.key_values:
                vals = ", ".join(escape_value(row[k]) for k in request.primary_keys)
                tuples.append(f"({vals})")
            where_clause = f"({pk_cols}) IN ({', '.join(tuples)})"

        sql = f"DELETE FROM {table_quoted} WHERE {where_clause}"

        logger.info(
            f"批量删除: 表={table_name}, 主键={request.primary_keys}, 行数={len(request.key_values)}"
        )

        try:
            result = SQLExecutor.execute(engine_key, config_dict, sql)
            elapsed = (datetime.now() - start_time).total_seconds() * 1000

            return BatchDeleteResult(
                success=result.success,
                deleted_count=result.affected_rows or 0,
                failed_count=0 if result.success else len(request.key_values),
                error_message=result.error_message,
                execution_time_ms=elapsed,
            )
        except Exception as e:
            elapsed = (datetime.now() - start_time).total_seconds() * 1000
            logger.error(f"批量删除执行失败: {e}")
            return BatchDeleteResult(
                success=False,
                deleted_count=0,
                failed_count=len(request.key_values),
                error_message=str(e),
                execution_time_ms=elapsed,
            )

    @staticmethod
    def insert_row(
        user_id: str, config_id: str, table_name: str, request: "InsertRowRequest"
    ) -> "RowOperationResult":
        from app.models.database_tool_models import (
            InsertRowRequest,
            RowOperationResult,
        )

        start_time = datetime.now()

        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            return RowOperationResult(
                success=False,
                affected_rows=0,
                error_message="Configuration not found",
                execution_time_ms=0,
            )

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            return RowOperationResult(
                success=False,
                affected_rows=0,
                error_message="Failed to decrypt password",
                execution_time_ms=0,
            )

        target_db = (
            request.database_name
            if request.database_name
            else config_row["database_name"]
        )

        target_schema = getattr(request, 'schema_name', None)

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": target_db,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
            "max_pool_size": config_row["max_pool_size"],
        }

        engine_key = f"{config_id}:{target_db}" if request.database_name else config_id
        db_type = config_row["db_type"]

        def quote_ident(name: str) -> str:
            if db_type == DatabaseType.SQLSERVER:
                return f"[{name}]"
            if db_type == DatabaseType.POSTGRESQL:
                return f'"{name}"'
            return f"`{name}`"

        def escape_value(val: Any) -> str:
            if val is None:
                return "NULL"
            if isinstance(val, bool):
                if db_type == DatabaseType.POSTGRESQL:
                    return "TRUE" if val else "FALSE"
                return "1" if val else "0"
            if isinstance(val, (int, float)):
                return str(val)
            escaped = str(val).replace("'", "''")
            return f"'{escaped}'"

        if target_schema and db_type == DatabaseType.POSTGRESQL:
            table_quoted = f'{quote_ident(target_schema)}.{quote_ident(table_name)}'
        else:
            table_quoted = quote_ident(table_name)
        col_names = ", ".join(quote_ident(k) for k in request.columns.keys())
        col_values = ", ".join(escape_value(v) for v in request.columns.values())

        sql = f"INSERT INTO {table_quoted} ({col_names}) VALUES ({col_values})"

        logger.info(f"插入行: 表={table_name}, 列={list(request.columns.keys())}")

        try:
            result = SQLExecutor.execute(engine_key, config_dict, sql)
            elapsed = (datetime.now() - start_time).total_seconds() * 1000

            return RowOperationResult(
                success=result.success,
                affected_rows=result.affected_rows or 0,
                execution_time_ms=elapsed,
                error_message=result.error_message,
            )
        except Exception as e:
            elapsed = (datetime.now() - start_time).total_seconds() * 1000
            logger.error(f"插入行失败: {e}")
            return RowOperationResult(
                success=False,
                affected_rows=0,
                error_message=str(e),
                execution_time_ms=elapsed,
            )

    @staticmethod
    def update_row(
        user_id: str, config_id: str, table_name: str, request: "UpdateRowRequest"
    ) -> "RowOperationResult":
        from app.models.database_tool_models import (
            UpdateRowRequest,
            RowOperationResult,
        )

        start_time = datetime.now()

        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            return RowOperationResult(
                success=False,
                affected_rows=0,
                error_message="Configuration not found",
                execution_time_ms=0,
            )

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            return RowOperationResult(
                success=False,
                affected_rows=0,
                error_message="Failed to decrypt password",
                execution_time_ms=0,
            )

        if not request.primary_keys or len(request.primary_keys) == 0:
            return RowOperationResult(
                success=False,
                affected_rows=0,
                error_message="Primary key is required for update",
                execution_time_ms=0,
            )

        target_db = (
            request.database_name
            if request.database_name
            else config_row["database_name"]
        )

        target_schema = getattr(request, 'schema_name', None)

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": target_db,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
            "max_pool_size": config_row["max_pool_size"],
        }

        engine_key = f"{config_id}:{target_db}" if request.database_name else config_id
        db_type = config_row["db_type"]

        def quote_ident(name: str) -> str:
            if db_type == DatabaseType.SQLSERVER:
                return f"[{name}]"
            if db_type == DatabaseType.POSTGRESQL:
                return f'"{name}"'
            return f"`{name}`"

        def escape_value(val: Any) -> str:
            if val is None:
                return "NULL"
            if isinstance(val, bool):
                if db_type == DatabaseType.POSTGRESQL:
                    return "TRUE" if val else "FALSE"
                return "1" if val else "0"
            if isinstance(val, (int, float)):
                return str(val)
            escaped = str(val).replace("'", "''")
            return f"'{escaped}'"

        if target_schema and db_type == DatabaseType.POSTGRESQL:
            table_quoted = f'{quote_ident(target_schema)}.{quote_ident(table_name)}'
        else:
            table_quoted = quote_ident(table_name)

        set_parts = []
        for col_name, col_value in request.columns.items():
            set_parts.append(f"{quote_ident(col_name)} = {escape_value(col_value)}")
        set_clause = ", ".join(set_parts)

        where_parts = []
        for pk in request.primary_keys:
            pk_value = request.key_values.get(pk)
            where_parts.append(f"{quote_ident(pk)} = {escape_value(pk_value)}")
        where_clause = " AND ".join(where_parts)

        sql = f"UPDATE {table_quoted} SET {set_clause} WHERE {where_clause}"

        logger.info(
            f"更新行: 表={table_name}, 主键={request.primary_keys}, 列={list(request.columns.keys())}"
        )

        try:
            result = SQLExecutor.execute(engine_key, config_dict, sql)
            elapsed = (datetime.now() - start_time).total_seconds() * 1000

            return RowOperationResult(
                success=result.success,
                affected_rows=result.affected_rows or 0,
                execution_time_ms=elapsed,
                error_message=result.error_message,
            )
        except Exception as e:
            elapsed = (datetime.now() - start_time).total_seconds() * 1000
            logger.error(f"更新行失败: {e}")
            return RowOperationResult(
                success=False,
                affected_rows=0,
                error_message=str(e),
                execution_time_ms=elapsed,
            )

    # ============ 表详情与备份管理 ============

    @staticmethod
    def get_table_detail(
        user_id: str,
        config_id: str,
        table_name: str,
        database_name: Optional[str] = None,
        schema_name: Optional[str] = None,
    ) -> TableDetailResponse:
        """获取表详细结构（字段、索引、外键）"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        target_db = database_name if database_name else config_row["database_name"]
        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": target_db,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{target_db}" if database_name else config_id
        engine = DBConnectionManager.get_engine(engine_key, config_dict)
        inspector = inspect(engine)

        # PostgreSQL: 使用 schema 参数传递给 inspector 方法
        is_postgresql = config_row["db_type"] == DatabaseType.POSTGRESQL
        schema_kw = {"schema": schema_name} if (is_postgresql and schema_name) else {}

        # 表注释
        table_comment = None
        try:
            table_comment = inspector.get_table_comment(table_name, **schema_kw).get("text")
        except Exception:
            pass

        # 字段
        columns = []
        pk_constraint = inspector.get_pk_constraint(table_name, **schema_kw)
        pk_columns = set(pk_constraint.get("constrained_columns", [])) if pk_constraint else set()

        for idx, col in enumerate(inspector.get_columns(table_name, **schema_kw)):
            col_type = str(col["type"])
            length = None
            if "(" in col_type:
                import re
                match = re.search(r"\(([^)]+)\)", col_type)
                if match:
                    length = match.group(1)
                    col_type = col_type.split("(")[0]

            columns.append(
                ColumnDetail(
                    name=col["name"],
                    type=col_type,
                    length=length,
                    nullable=col.get("nullable", True),
                    default_value=str(col["default"]) if col.get("default") is not None else None,
                    comment=col.get("comment"),
                    primary_key=col["name"] in pk_columns,
                    auto_increment=col.get("autoincrement", False),
                    ordinal_position=idx + 1,
                )
            )

        # 索引
        indexes = []
        for idx in inspector.get_indexes(table_name, **schema_kw):
            indexes.append(
                IndexDetail(
                    name=idx["name"],
                    unique=idx.get("unique", False),
                    primary=idx["name"] == pk_constraint.get("name") if pk_constraint else False,
                    columns=idx.get("column_names", []),
                )
            )

        # 外键
        foreign_keys = []
        for fk in inspector.get_foreign_keys(table_name, **schema_kw):
            foreign_keys.append(
                ForeignKeyDetail(
                    name=fk.get("name", ""),
                    constrained_columns=fk.get("constrained_columns", []),
                    referred_table=fk.get("referred_table", ""),
                    referred_columns=fk.get("referred_columns", []),
                )
            )

        # 行数
        row_count = None
        try:
            with engine.connect() as conn:
                if is_postgresql and schema_name:
                    count_sql = text(f'SELECT COUNT(*) FROM "{schema_name}"."{table_name}"')
                else:
                    count_sql = text(f"SELECT COUNT(*) FROM \"{table_name}\"")
                result = conn.execute(count_sql)
                row_count = result.scalar()
        except Exception:
            pass

        return TableDetailResponse(
            table_name=table_name,
            comment=table_comment,
            columns=columns,
            indexes=indexes,
            foreign_keys=foreign_keys,
            row_count=row_count,
        )

    @staticmethod
    def get_table_row_count(
        user_id: str,
        config_id: str,
        table_name: str,
        database_name: str,
        schema_name: Optional[str] = None,
    ) -> int:
        """获取表行数"""
        config_row = DatabaseToolService._get_config_with_password(config_id, user_id)
        if not config_row:
            raise ValueError("Configuration not found")

        try:
            password = EncryptionUtils.decrypt(config_row["password_encrypted"])
        except Exception:
            raise ValueError("Failed to decrypt password")

        config_dict = {
            "db_type": config_row["db_type"],
            "host": config_row["host"],
            "port": config_row["port"],
            "database_name": database_name,
            "username": config_row["username"],
            "password": password,
            "charset": config_row["charset"],
        }

        engine_key = f"{config_id}:{database_name}"
        engine = DBConnectionManager.get_engine(engine_key, config_dict)

        # PostgreSQL: use schema-qualified table name for COUNT query
        if config_row["db_type"] == DatabaseType.POSTGRESQL and schema_name:
            table_ref = f'"{schema_name}"."{table_name}"'
        else:
            table_ref = f'"{table_name}"'

        try:
            with engine.connect() as conn:
                result = conn.execute(text(f"SELECT COUNT(*) FROM {table_ref}"))
                return result.scalar() or 0
        except Exception as e:
            logger.warning(f"Failed to get row count for {table_name}: {e}")
            return 0

    @staticmethod
    def list_backups(
        user_id: str,
        config_id: Optional[str] = None,
        database_name: Optional[str] = None,
        page: int = 1,
        page_size: int = 20,
    ) -> Dict[str, Any]:
        """获取备份历史列表"""
        return backup_storage.list_records(
            user_id=user_id,
            config_id=config_id,
            database_name=database_name,
            page=page,
            page_size=page_size,
        )

    @staticmethod
    def delete_backup(user_id: str, backup_id: str) -> bool:
        """删除备份"""
        return backup_storage.delete_record(backup_id, user_id)

    @staticmethod
    def get_backup_file_path(user_id: str, backup_id: str) -> Optional[str]:
        """获取备份文件路径（用于下载）"""
        record = backup_storage.get_record(backup_id, user_id)
        if record and record.file_path:
            return record.file_path
        return None

    @staticmethod
    def increment_download_count(user_id: str, backup_id: str) -> bool:
        """增加备份下载计数"""
        return backup_storage.increment_download_count(backup_id, user_id)

    @staticmethod
    def get_display_preferences(user_id: str) -> DisplayPreferenceResponse:
        """获取用户的显示偏好"""
        conn = get_pooled_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT visible_connections, visible_databases, updated_at "
                    "FROM user_display_preferences WHERE user_id = %s",
                    (user_id,),
                )
                row = cur.fetchone()
                if row:
                    return DisplayPreferenceResponse(
                        visible_connections=row["visible_connections"],
                        visible_databases=row["visible_databases"],
                        updated_at=row["updated_at"],
                    )
                return DisplayPreferenceResponse()
        finally:
            release_db_connection(conn)

    @staticmethod
    def save_display_preferences(
        user_id: str, preferences: dict
    ) -> DisplayPreferenceResponse:
        """保存用户的显示偏好"""
        import json as json_mod

        conn = get_pooled_db_connection()
        try:
            visible_conn = preferences.get("visible_connections")
            visible_db = preferences.get("visible_databases")
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO user_display_preferences "
                    "(user_id, visible_connections, visible_databases) "
                    "VALUES (%s, %s, %s) "
                    "ON CONFLICT (user_id) DO UPDATE SET "
                    "visible_connections = EXCLUDED.visible_connections, "
                    "visible_databases = EXCLUDED.visible_databases, "
                    "updated_at = CURRENT_TIMESTAMP "
                    "RETURNING visible_connections, visible_databases, updated_at",
                    (
                        user_id,
                        json_mod.dumps(visible_conn) if visible_conn is not None else None,
                        json_mod.dumps(visible_db) if visible_db else None,
                    ),
                )
                row = cur.fetchone()
                conn.commit()
                return DisplayPreferenceResponse(
                    visible_connections=row["visible_connections"],
                    visible_databases=row["visible_databases"],
                    updated_at=row["updated_at"],
                )
        except Exception:
            conn.rollback()
            raise
        finally:
            release_db_connection(conn)
